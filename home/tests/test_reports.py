"""صورتحساب و خروجیِ اکسل."""
from io import BytesIO

from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from .factories import make_customer, make_owner, make_transaction


class StatementTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.owner = make_owner(fullname="مالکِ آزمایشی")

        cls.debtor = make_customer(cls.owner, "الف بدهکار", phone="09121110001")
        make_transaction(cls.owner, cls.debtor, debt=900_000)
        make_transaction(cls.owner, cls.debtor, paid=400_000)

        cls.creditor = make_customer(cls.owner, "ب بستانکار", phone="09121110002")
        make_transaction(cls.owner, cls.creditor, paid=250_000)

        cls.untouched = make_customer(cls.owner, "ج بی‌تراکنش", phone="09121110003")

    def setUp(self):
        self.client.force_login(self.owner)

    def body(self):
        return self.client.get(reverse("api:statement")).json()

    def test_every_customer_appears_including_those_without_transactions(self):
        rows = self.body()["rows"]
        self.assertEqual(len(rows), 3)
        self.assertIn("ج بی‌تراکنش", [row["fullname"] for row in rows])

    def test_rows_are_sorted_by_name(self):
        names = [row["fullname"] for row in self.body()["rows"]]
        self.assertEqual(names, sorted(names))

    def test_balance_and_status_per_row(self):
        rows = {row["fullname"]: row for row in self.body()["rows"]}
        self.assertEqual(rows["الف بدهکار"]["balance"], -500_000)
        self.assertEqual(rows["الف بدهکار"]["status"], "بدهکار")
        self.assertEqual(rows["ب بستانکار"]["status"], "بستانکار")
        self.assertEqual(rows["ج بی‌تراکنش"]["status"], "بی حساب")

    def test_totals_match_the_rows(self):
        body = self.body()
        rows = body["rows"]
        totals = body["totals"]
        self.assertEqual(totals["debt"], sum(row["debt"] for row in rows))
        self.assertEqual(totals["paid"], sum(row["paid"] for row in rows))
        self.assertEqual(totals["balance"], totals["paid"] - totals["debt"])
        self.assertEqual(totals["customers"], len(rows))
        self.assertEqual(totals["debtors"], 1)
        self.assertEqual(totals["creditors"], 1)

    def test_owner_and_date_are_included_for_the_letterhead(self):
        body = self.body()
        self.assertEqual(body["owner"]["fullname"], "مالکِ آزمایشی")
        self.assertIn("month_label", body["generated_at"])

    def test_statement_is_not_paginated(self):
        """صورتحسابِ نصفه صورتحساب نیست."""
        for index in range(10):
            make_customer(self.owner, f"مشتری {index}")
        body = self.body()
        self.assertEqual(len(body["rows"]), 13)
        self.assertNotIn("next", body)

    def test_another_owners_customers_are_absent(self):
        other = make_owner()
        make_customer(other, "مشتریِ دیگری", phone="09129990000")
        make_transaction(other, make_customer(other, "دومی", phone="09129990001"), debt=5_000_000)

        body = self.body()
        self.assertNotIn("مشتریِ دیگری", [row["fullname"] for row in body["rows"]])
        self.assertEqual(body["totals"]["customers"], 3)

    def test_anonymous_is_refused(self):
        self.client.logout()
        self.assertIn(self.client.get(reverse("api:statement")).status_code,
                      (status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN))


class ExcelExportTests(APITestCase):
    def setUp(self):
        self.owner = make_owner()
        self.client.force_login(self.owner)
        self.customer = make_customer(self.owner, "مشتری", phone="09121110001")
        make_transaction(self.owner, self.customer, debt=700_000, jalali=(1405, 2, 19))
        make_transaction(self.owner, self.customer, paid=300_000, jalali=(1405, 3, 1))

    def download(self):
        response = self.client.get(reverse("api:excel_export"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        return response

    def workbook(self, response):
        return load_workbook(BytesIO(b"".join(response.streaming_content)))

    def test_it_downloads_as_an_xlsx_attachment(self):
        response = self.download()
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])

    def test_it_has_both_sheets(self):
        workbook = self.workbook(self.download())
        self.assertEqual(workbook.sheetnames, ["مشتریان", "تراکنش ها"])

    def test_sheets_are_right_to_left(self):
        """بدونِ این، اکسل ستون‌ها را از چپ می‌چیند و جدولِ فارسی وارونه می‌شود."""
        workbook = self.workbook(self.download())
        for name in workbook.sheetnames:
            self.assertTrue(workbook[name].sheet_view.rightToLeft, msg=name)

    def test_customer_sheet_carries_the_totals(self):
        sheet = self.workbook(self.download())["مشتریان"]
        # سرستون + یک مشتری + ردیفِ جمع
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet.cell(row=2, column=3).value, 700_000)
        self.assertEqual(sheet.cell(row=2, column=5).value, -400_000)
        self.assertEqual(sheet.cell(row=3, column=1).value, "جمع کل")
        self.assertEqual(sheet.cell(row=3, column=3).value, 700_000)

    def test_transaction_sheet_lists_every_row_with_a_jalali_date(self):
        sheet = self.workbook(self.download())["تراکنش ها"]
        self.assertEqual(sheet.max_row, 4)   # سرستون + دو تراکنش + جمع
        dates = {sheet.cell(row=row, column=5).value for row in (2, 3)}
        self.assertEqual(dates, {"1405/02/19", "1405/03/01"})

    def test_another_owners_rows_are_absent(self):
        other = make_owner()
        other_customer = make_customer(other, "مشتریِ دیگری", phone="09129990000")
        make_transaction(other, other_customer, debt=9_000_000)

        sheet = self.workbook(self.download())["مشتریان"]
        names = {sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)}
        self.assertNotIn("مشتریِ دیگری", names)

    def test_anonymous_is_refused(self):
        self.client.logout()
        self.assertIn(self.client.get(reverse("api:excel_export")).status_code,
                      (status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN))
