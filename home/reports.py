"""گزارش‌ها — صورتحسابِ همهٔ مشتریان و خروجیِ اکسل.

جدا از `dashboard.py` است چون کارِ متفاوتی می‌کند: آنجا چند عددِ خلاصه برای دیدن
روی صفحه ساخته می‌شود، اینجا **کلِ دفتر** برای بیرون بردن — چاپ یا فایل. به همین
دلیل هم صفحه‌بندی ندارد: صورتحسابِ نصفه صورتحساب نیست.

⚠️ مثل هر جای دیگر، همه‌چیز به مالکِ درخواست محدود است و وضعیتِ حساب از
تراکنش‌های همین مالک حساب می‌شود، نه از ستونِ کش‌شدهٔ `Customer.code`.
"""
from io import BytesIO

import jdatetime
from django.db.models import Count, Q, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Customer, CustomerOwner, Transaction

STATUS_LABELS = {1: "بستانکار", 0: "بی حساب", -1: "بدهکار"}

JALALI_MONTHS = [
    "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند",
]


def _today():
    today = jdatetime.date.fromgregorian(date=timezone.localtime().date())
    return {"year": today.year, "month": today.month, "day": today.day,
            "month_label": JALALI_MONTHS[today.month - 1]}


def build_statement(user):
    """صورتحسابِ همهٔ مشتریانِ یک مالک — هر مشتری با مجموع‌ها و مانده‌اش.

    یک کوئریِ گروهی برای همه، نه یکی به‌ازای هر مشتری. زیرکوئریِ شناسه به‌جای
    JOIN روی `owners` استفاده می‌شود چون آن یکی کنارِ aggregate ردیف‌ها را تکثیر
    می‌کند و مجموع‌ها را بزرگ نشان می‌دهد.
    """
    owned = CustomerOwner.objects.filter(owner=user).values("customer_id")
    mine = Q(customer_transactions__owner=user)

    customers = (Customer.objects
                 .filter(id__in=owned)
                 .annotate(
                     total_debt=Coalesce(Sum("customer_transactions__debt", filter=mine), 0),
                     total_paid=Coalesce(Sum("customer_transactions__paid", filter=mine), 0),
                     tx_count=Count("customer_transactions", filter=mine),
                 )
                 .order_by("fullname"))

    rows = []
    totals = {"debt": 0, "paid": 0, "transactions": 0, "debtors": 0, "creditors": 0}
    for customer in customers:
        balance = customer.total_paid - customer.total_debt
        code = 1 if balance > 0 else (-1 if balance < 0 else 0)
        rows.append({
            "id": customer.id,
            "fullname": customer.fullname,
            "phone": customer.phone,
            "debt": customer.total_debt,
            "paid": customer.total_paid,
            "balance": balance,
            "code": code,
            "status": STATUS_LABELS[code],
            "transactions": customer.tx_count,
        })
        totals["debt"] += customer.total_debt
        totals["paid"] += customer.total_paid
        totals["transactions"] += customer.tx_count
        if code == -1:
            totals["debtors"] += 1
        elif code == 1:
            totals["creditors"] += 1

    totals["balance"] = totals["paid"] - totals["debt"]
    totals["customers"] = len(rows)

    return {
        "owner": {"fullname": user.fullname, "phone": user.phone},
        "generated_at": _today(),
        "rows": rows,
        "totals": totals,
    }


# ------------------------------------------------------------ خروجیِ اکسل

_HEADER_FILL = PatternFill("solid", fgColor="0D121E")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="EDF2F7")
_TOTAL_FONT = Font(bold=True, size=11)
_THIN = Side(style="thin", color="D8DEE7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0"


def _write_sheet(sheet, headers, widths, rows, total_row=None):
    """یک برگهٔ آماده: سرستونِ تیره، بوردر، عرضِ ستون و ردیفِ جمع.

    ⚠️ `sheet_view.rightToLeft` لازم است: بدونش اکسل ستون‌ها را از چپ می‌چیند و
    یک جدولِ فارسی وارونه به نظر می‌رسد.
    """
    sheet.sheet_view.rightToLeft = True
    sheet.append(headers)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    sheet.row_dimensions[1].height = 26

    for row in rows:
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, int):
                cell.number_format = _MONEY

    if total_row:
        sheet.append(total_row)
        for cell in sheet[sheet.max_row]:
            cell.fill = _TOTAL_FILL
            cell.font = _TOTAL_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, int):
                cell.number_format = _MONEY

    # سرستون هنگام اسکرول ثابت بماند — در فهرستِ بلند بدونش نمی‌شود فهمید
    # هر ستون چیست
    sheet.freeze_panes = "A2"


def build_workbook(user):
    """دفترِ کاملِ مالک در یک فایلِ اکسل، با دو برگه.

    این خروجی به‌جای «گزارش»، نقشِ **پشتیبانِ خودِ کاربر** را دارد: اگر روزی
    سرور از دست برود، دفتر از دست نرفته. پس همهٔ ردیف‌ها می‌آیند، نه خلاصه.
    """
    statement = build_statement(user)
    workbook = Workbook()

    # ── برگهٔ مشتریان ──────────────────────────────────────────────────
    customers_sheet = workbook.active
    customers_sheet.title = "مشتریان"
    _write_sheet(
        customers_sheet,
        ["نام و نام خانوادگی", "شماره تماس", "مجموع نسیه", "مجموع پرداختی", "مانده", "وضعیت", "تعداد تراکنش"],
        [28, 16, 16, 16, 16, 12, 14],
        [[row["fullname"], row["phone"], row["debt"], row["paid"], row["balance"],
          row["status"], row["transactions"]] for row in statement["rows"]],
        total_row=["جمع کل", "", statement["totals"]["debt"], statement["totals"]["paid"],
                   statement["totals"]["balance"], "", statement["totals"]["transactions"]],
    )

    # ── برگهٔ تراکنش‌ها ────────────────────────────────────────────────
    transactions = (Transaction.objects
                    .filter(owner=user)
                    .select_related("customer")
                    .order_by("-created", "-id"))
    transactions_sheet = workbook.create_sheet("تراکنش ها")
    _write_sheet(
        transactions_sheet,
        ["مشتری", "شماره تماس", "نسیه", "پرداختی", "تاریخ"],
        [28, 16, 16, 16, 14],
        [[t.customer.fullname, t.customer.phone, t.debt, t.paid,
          f"{t.year}/{t.month:02d}/{t.day:02d}" if t.year else ""]
         for t in transactions],
        total_row=["جمع کل", "", statement["totals"]["debt"], statement["totals"]["paid"], ""],
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    today = statement["generated_at"]
    filename = f"crm-{today['year']}-{today['month']:02d}-{today['day']:02d}.xlsx"
    return stream, filename
